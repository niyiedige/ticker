import openpyxl
# to modifide the output :1 change the title here 2.change the source of the output in getinfo
def output(output):

    title = [ 'Ticker', 'Quote','Numberchange','Percentagechange']
    wb=openpyxl.load_workbook('result.xlsx')
    try :
        #depend on the frequency
        ws1=wb.get_sheet_by_name("ticker")
        wb.remove_sheet(ws1)
    except KeyError:
        pass
    ws=wb.create_sheet("ticker")
    for col in ws.iter_cols(min_row=1, max_col=4):
        for cell in col:
            cell.value=title.pop()
    rownum=1

    while output:
        colnum = 0
        rownum=rownum+1
    #    print(output)
        job=output.pop()
    #    print(1)
        while job:
            colnum+=1
            a = job.pop()
            ws.cell(row=rownum,column=colnum).value = a



    wb.save("result.xlsx")
